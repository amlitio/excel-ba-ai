import streamlit as st
import pandas as pd
import openpyxl as px
import openai

# OpenAI API Key
openai_api_key = "YOUR_OPENAI_API_KEY"

def review_excel(file):
    # Read the Excel file and perform a basic review
    df = pd.read_excel(file)
    return df.describe()

def check_formulas(file):
    # Check for formula errors in the Excel file
    workbook = px.load_workbook(file)
    sheet = workbook.active
    feedback = []
    # Example: Check for specific formulas and provide feedback
    return feedback

def suggest_python_integration(file):
    workbook = px.load_workbook(file)
    sheet = workbook.active
    suggestions = []

    # Analyzing existing formulas and data
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                formula = cell.value[1:]
                suggestion = f"For the formula '{formula}' in cell {cell.coordinate}, you can use Python's Pandas library to perform the calculation."
                suggestions.append(suggestion)

    # Adding general suggestions for business analysis tasks
    suggestions.append("For data cleaning and preparation, you can use Python's Pandas library to remove outliers, impute missing values, and normalize data.")
    suggestions.append("For statistical analysis, you can use Python's SciPy library to perform regression analysis, clustering, and classification.")
    suggestions.append("For machine learning, you can use Python's scikit-learn library to build models like decision trees, random forests, and support vector machines.")
    suggestions.append("For natural language processing, you can use Python's NLTK library to perform sentiment analysis, topic modeling, and named entity recognition.")
    suggestions.append("For computer vision, you can use Python's OpenCV library to perform object detection, image classification, and face recognition.")

    return suggestions

def get_openai_feedback(file):
    # Use OpenAI to provide feedback on Excel skills
    prompt = f"Review the uploaded Excel file and provide suggestions on how to improve Excel skills. File: {file.name}"
    response = openai.Completion.create(
        engine="text-davinci-003",
        prompt=prompt,
        max_tokens=150
    )
    return response.choices[0].text

st.title("Excel Business Analysis Tool")
uploaded_file = st.file_uploader("Choose an Excel file", type="xlsx")

if uploaded_file:
    st.write("Reviewing the Excel file...")
    analysis_result = review_excel(uploaded_file)
    st.subheader("Analysis Results:")
    st.write(analysis_result)

    st.subheader("Formula Check:")
    formula_feedback = check_formulas(uploaded_file)
    for item in formula_feedback:
        st.write(item)

    st.subheader("Python Integration Suggestions:")
    python_suggestions = suggest_python_integration(uploaded_file)
    for suggestion in python_suggestions:
        st.write(suggestion)

    st.subheader("OpenAI Feedback on Excel Skills:")
    openai_feedback = get_openai_feedback(uploaded_file)
    st.write(openai_feedback)
